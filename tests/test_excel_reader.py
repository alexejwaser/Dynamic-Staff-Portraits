# tests/test_excel_reader.py
from pathlib import Path
from app.core.excel.reader import ExcelReader
import openpyxl

def create_sample(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Standort1'
    ws['A1'] = 'Klasse'
    ws['B1'] = 'Nachname'
    ws['C1'] = 'Vorname'
    ws['D1'] = 'SchuelerID'
    ws['E1'] = 'Fotografiert?'
    ws['F1'] = 'Aufnahmedatum'
    ws.append(['INF', 'Meier', 'Hans', '001', '', ''])
    ws.append(['INF', 'Muster', 'Eva', '002', '', ''])
    wb.save(path)

def test_read(tmp_path):
    xl = tmp_path / 'test.xlsx'
    create_sample(xl)
    mapping = {
        'klasse':'A', 'nachname':'B', 'vorname':'C', 'schuelerId':'D',
        'fotografiert':'E', 'aufnahmedatum':'F'
    }
    reader = ExcelReader(xl, mapping)
    classes = reader.classes_for_location('Standort1')
    assert classes == ['INF']
    learners = reader.learners('Standort1', 'INF')
    assert len(learners) == 2
    assert learners[0].nachname == 'Meier'
    # mark photographed and skipped
    reader.mark_photographed('Standort1', learners[0].row, True, '01.01.2024')
    reader.mark_photographed('Standort1', learners[1].row, False)
    wb = openpyxl.load_workbook(xl)
    ws = wb['Standort1']
    assert ws['E2'].value == 'Ja'
    assert ws['F2'].value == '01.01.2024'
    assert ws['E3'].value == 'Nein'
    assert ws['F3'].value is None
