# app/core/excel/reader.py
from dataclasses import dataclass
from typing import List
from pathlib import Path
import openpyxl
@dataclass
class Learner:
    klasse: str
    nachname: str
    vorname: str
    schueler_id: str = ''
    row: int = 0
    is_new: bool = False

class ExcelReader:
    def __init__(self, path: Path, mapping: dict):
        self.path = path
        self.mapping = mapping
        self.wb = openpyxl.load_workbook(path)

    def locations(self) -> List[str]:
        return self.wb.sheetnames

    def classes_for_location(self, location: str) -> List[str]:
        sheet = self.wb[location]
        col = self.mapping['klasse']
        values = {str(cell.value) for cell in sheet[col][1:] if cell.value}
        return sorted(values)

    def learners(self, location: str, class_name: str) -> List[Learner]:
        sheet = self.wb[location]
        m = self.mapping
        learners = []
        for row in sheet.iter_rows(min_row=2):
            if str(row[openpyxl.utils.column_index_from_string(m['klasse'])-1].value) == class_name:
                nachname = row[openpyxl.utils.column_index_from_string(m['nachname'])-1].value
                vorname = row[openpyxl.utils.column_index_from_string(m['vorname'])-1].value
                sid = row[openpyxl.utils.column_index_from_string(m['schuelerId'])-1].value
                if nachname and vorname and sid:
                    learners.append(Learner(str(class_name), str(nachname), str(vorname), str(sid), row=row[0].row))
        learners.sort(key=lambda l: (l.nachname, l.vorname))
        return learners

    def mark_photographed(self, location: str, row: int, photographed: bool, date: str | None = None) -> None:
        sheet = self.wb[location]
        col_phot = self.mapping.get('fotografiert')
        col_date = self.mapping.get('aufnahmedatum')
        if col_phot:
            sheet[f"{col_phot}{row}"].value = 'Ja' if photographed else 'Nein'
        if col_date:
            sheet[f"{col_date}{row}"].value = date if photographed else None
        self.wb.save(self.path)
